import pandas as pd
from tkinter import filedialog
from openpyxl import load_workbook

class Model:

    @staticmethod
    def export_mobile_by_brand_data(Data_dict : dict, FileName : str = "MobileBrands_Data"):
        """
        It will export the mobile data scrapped by brands in an Excel sheet,
        It contains workbooks with name of Mobile brands that selected while
        scraping.
        :param Data_dict: (dict) -> The 2D Dictionary having mobile brands name and a list of their mobile name with prices
        :param FileName: (str) -> The File name in which the file to be exported
        :return: (str) -> Return the status "FILE SAVED" if file is saved successfully, else "FILEPATH is Empty"
        """

        Sheets = list(Data_dict.keys())
        for Sheet, Dict in Data_dict.items():
            Dataframe = pd.DataFrame(Dict)
            Dataframe.style.set_table_styles(
                [
                    {
                        'selector': 'th',
                        'props': [('font-weight', 'bold'), ('background-color', '#ADD8E6')]
                    }
                ]
            )
            Data_dict[Sheet] = Dataframe

        # Open the save file dialog
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 initialfile=FileName,
                                                 filetypes=[("EXCEL files", "*.xlsx")])

        if len(file_path.strip()) == 0:
            return "FILEPATH EMPTY"

        else:

            with pd.ExcelWriter(file_path, mode = 'w') as writer:
                pd.DataFrame().to_excel(writer, sheet_name = "EmptySheet", index = False)

            for Sheet in Sheets:
                with pd.ExcelWriter(file_path, if_sheet_exists = 'replace', engine = 'openpyxl', mode = 'a') as writer:
                    Data_dict[Sheet].to_excel(writer, sheet_name = Sheet)

            wb = load_workbook(file_path)
            if "EmptySheet" in wb.sheetnames:
                sheet = wb["EmptySheet"]
                sheet.sheet_state = 'hidden'
            wb.save(file_path)

            return "FILE SAVED"
    @staticmethod
    def export_mobile_by_name_data(Data_dict : dict, FileName : str):
        """
        It will export the individual mobile specs in detail in an excel file
        :param Data_dict: (dict) -> The dictionary with mobile features and its values
        :param FileName: (str) -> The File name in which the file to be exported
        :return: (str) -> Return the status "FILE SAVED" if file is saved successfully, else "FILEPATH is Empty"
        """

        df = pd.DataFrame(list(Data_dict.items()), columns=['Key', 'Value'])

        def highlight_first_row(row):
            return ['font-weight: bold; background-color: yellow' if row.name == 0 else '' for _ in row]
        styled_df = df.style.apply(highlight_first_row, axis=1)

        FileName = '_'.join(FileName.split()) + "_" + 'Specs'
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 initialfile=FileName,
                                                 filetypes=[("EXCEL files", "*.xlsx")])

        if len(file_path.strip()) == 0:
            return "FILEPATH EMPTY"

        else:
            # Export the styled DataFrame to an Excel file
            styled_df.to_excel(file_path, engine = 'openpyxl',index=False, header=False)

            return "FILE SAVED"
